from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from apps.productos.models import Producto, Categoria
from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField, FloatField
from django.db.models.functions import Coalesce
from django.db import models
import csv
from django.http import HttpResponse
from datetime import datetime
from django.core.paginator import Paginator
from apps.transacciones.models import Transaccion
from django.db.models import Sum, Q
from datetime import datetime, timedelta
from django.utils import timezone
import json
from django.contrib import messages
import io  # Add this import for StringIO
from django.shortcuts import redirect  # Also add this for the redirect function

@login_required
def reportes_view(request):
    """Vista principal del panel de reportes"""
    return render(request, 'reportes/reportes.html')

# Reportes Financieros y Fiscales
@login_required
def ventas_report(request):
    """Reporte de ventas diarias, semanales, mensuales y anuales"""
    return render(request, 'reportes/ventas.html')

@login_required
def iva_report(request):
    """Reporte de IVA"""
    return render(request, 'reportes/iva.html')

@login_required
def facturas_report(request):
    """Reporte de facturas y comprobantes"""
    return render(request, 'reportes/facturas.html')

# Reportes Operativos
@login_required
def inventario_report(request):
    """Reporte de inventario mejorado - Nivel ERP Empresarial"""
    from decimal import Decimal
    
    # 🎯 Obtener datos del negocio del usuario
    try:
        from apps.usuarios.models import Business
        business = Business.objects.get(user=request.user)
        nombre_negocio = business.nombre_negocio or request.user.nombre_completo
        ruc_negocio = business.ruc_negocio or 'N/A'
    except Business.DoesNotExist:
        nombre_negocio = request.user.nombre_completo
        ruc_negocio = request.user.ruc_negocio if hasattr(request.user, 'ruc_negocio') else 'N/A'
    
    # 🎯 CAMBIO 1: Filtrar SOLO productos inventariables (físicos e insumos)
    # NO mostrar servicios ni combos sin inventario
    productos = Producto.objects.filter(
        usuario_creador=request.user,
        activo=True,
        tipo_producto__in=['fisico', 'insumo'],  # Solo inventariables
        controla_stock=True  # Solo los que controlan stock
    ).order_by('stock')
    
    # 🎯 CAMBIO 2: Stock bajo usando stock_minimo (dinámico, no fijo)
    # Ahora es: stock_actual <= stock_minimo
    productos_stock_bajo = productos.filter(
        stock__lte=models.F('stock_minimo'),
        stock__gt=0
    )
    
    # Productos agotados (solo los que controlan stock y tienen 0)
    productos_agotados = productos.filter(stock=0)
    
    # Productos por categoría - Solo categorías con productos inventariables
    categorias = Categoria.objects.filter(
        producto__usuario_creador=request.user,
        producto__tipo_producto__in=['fisico', 'insumo'],
        producto__controla_stock=True
    ).distinct().annotate(
        total_productos=Count('producto', filter=Q(
            producto__usuario_creador=request.user,
            producto__tipo_producto__in=['fisico', 'insumo'],
            producto__controla_stock=True
        ))
    )
    
    # Calculate valor_inventario separately for each category
    for categoria in categorias:
        productos_categoria = Producto.objects.filter(
            categoria=categoria, 
            usuario_creador=request.user,
            activo=True,
            tipo_producto__in=['fisico', 'insumo'],
            controla_stock=True
        )
        # Calcular valor basado en costo (contablemente correcto)
        valor_inventario_costo = sum(
            (p.costo or Decimal('0')) * Decimal(str(p.stock or 0)) 
            for p in productos_categoria if p.stock is not None
        )
        # Calcular valor basado en precio de venta
        valor_inventario_venta = sum(
            p.precio * Decimal(str(p.stock or 0)) 
            for p in productos_categoria if p.stock is not None
        )
        categoria.valor_inventario_costo = valor_inventario_costo
        categoria.valor_inventario_venta = valor_inventario_venta
        categoria.utilidad_potencial = valor_inventario_venta - valor_inventario_costo
    
    # Valor total del inventario basado en COSTO (contablemente correcto)
    valor_total_costo = sum(
        (p.costo or Decimal('0')) * Decimal(str(p.stock or 0)) 
        for p in productos if p.stock is not None
    )
    
    # Valor total del inventario basado en PRECIO DE VENTA
    valor_total_venta = sum(
        p.precio * Decimal(str(p.stock or 0)) 
        for p in productos if p.stock is not None
    )
    
    # 🎯 CAMBIO 3: Valor en riesgo más preciso (stock bajo × costo, no precio)
    # Esto es el capital real en riesgo
    valor_en_riesgo = sum(
        (p.costo or Decimal('0')) * Decimal(str(p.stock or 0)) 
        for p in productos_stock_bajo if p.stock is not None
    )
    
    # Utilidad potencial (diferencia entre precio venta y costo)
    utilidad_potencial = valor_total_venta - valor_total_costo
    
    # 🎯 CAMBIO 4: Separar utilidad por tipo
    # Productos físicos
    productos_fisicos = productos.filter(tipo_producto='fisico')
    utilidad_fisicos = sum(
        (p.precio - (p.costo or Decimal('0'))) * Decimal(str(p.stock or 0))
        for p in productos_fisicos if p.stock is not None
    )
    
    # Insumos
    productos_insumos = productos.filter(tipo_producto='insumo')
    utilidad_insumos = sum(
        (p.precio - (p.costo or Decimal('0'))) * Decimal(str(p.stock or 0))
        for p in productos_insumos if p.stock is not None
    )
    
    # Promedio de valor por producto
    promedio_valor = valor_total_costo / productos.count() if productos.count() > 0 else Decimal('0')
    
    # Check if export was requested
    export_format = request.GET.get('export')
    if export_format in ['excel', 'pdf']:
        if export_format == 'excel':
            return exportar_inventario_excel(
                productos, categorias, valor_total_costo, valor_total_venta,
                productos_stock_bajo, productos_agotados, valor_en_riesgo,
                utilidad_potencial, promedio_valor, request.user
            )
        elif export_format == 'pdf':
            return exportar_inventario_pdf(
                productos, categorias, valor_total_costo, valor_total_venta,
                productos_stock_bajo, productos_agotados, valor_en_riesgo,
                utilidad_potencial, request.user
            )
    
    context = {
        'productos': productos,
        'productos_stock_bajo': productos_stock_bajo,
        'productos_agotados': productos_agotados,
        'categorias': categorias,
        'valor_total_costo': valor_total_costo,
        'valor_total_venta': valor_total_venta,
        'valor_en_riesgo': valor_en_riesgo,
        'utilidad_potencial': utilidad_potencial,
        'promedio_valor': promedio_valor,
        # 🎯 Nuevos indicadores profesionales
        'total_productos_fisicos': productos_fisicos.count(),
        'total_productos_insumos': productos_insumos.count(),
        'utilidad_fisicos': utilidad_fisicos,
        'utilidad_insumos': utilidad_insumos,
        'total_productos': productos.count(),
        'total_categorias': categorias.count(),
        'count_stock_bajo': productos_stock_bajo.count(),
        'count_agotados': productos_agotados.count(),
    }
    
    return render(request, 'reportes/inventario.html', context)

@login_required
def analisis_ventas_report(request):
    """Análisis de ventas"""
    return render(request, 'reportes/analisis_ventas.html')

@login_required
def clientes_report(request):
    """Reporte de clientes"""
    return render(request, 'reportes/clientes.html')

# Reportes para el SRI
@login_required
def ats_report(request):
    """Anexo Transaccional Simplificado (ATS)"""
    return render(request, 'reportes/ats.html')

@login_required
def iva_declaracion_report(request):
    """Declaración del IVA (Formulario 104)"""
    return render(request, 'reportes/iva_declaracion.html')

@login_required
def renta_report(request):
    """Declaración del Impuesto a la Renta"""
    return render(request, 'reportes/renta.html')

# Reportes Rápidos
@login_required
def ventas_diarias_report(request):
    """Reporte de ventas del día"""
    return render(request, 'reportes/ventas_diarias.html')

@login_required
def productos_top_report(request):
    """Reporte de productos más vendidos"""
    return render(request, 'reportes/productos_top.html')

@login_required
def stock_bajo_report(request):
    """Reporte de productos con stock bajo"""
    return render(request, 'reportes/stock_bajo.html')

@login_required
def generar_xml(request):
    """Generación de archivos XML para el SRI"""
    return render(request, 'reportes/generar_xml.html')

# Add this function to your views.py
def exportar_ventas(request):
    """
    Export sales data to CSV based on filter parameters
    """
    # Get filter parameters
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    tipo_documento = request.GET.get('tipo_documento', 'todos')
    
    # Convert string dates to datetime objects if provided
    if fecha_desde:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
    else:
        fecha_desde = datetime.now().date().replace(day=1)  # First day of current month
        
    if fecha_hasta:
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    else:
        fecha_hasta = datetime.now().date()
    
    # Query your database for facturas based on filters
    # This is a placeholder - adjust according to your models
    facturas = Factura.objects.filter(fecha__range=[fecha_desde, fecha_hasta])
    
    if tipo_documento != 'todos':
        facturas = facturas.filter(tipo=tipo_documento)
    
    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="ventas_{fecha_desde}_a_{fecha_hasta}.csv"'
    
    # Create CSV writer
    writer = csv.writer(response)
    writer.writerow(['Número', 'Fecha', 'Cliente', 'Subtotal', 'IVA', 'Total', 'Estado'])
    
    # Add data rows
    for factura in facturas:
        writer.writerow([
            factura.numero,
            factura.fecha.strftime('%d/%m/%Y'),
            factura.cliente.nombre,
            factura.subtotal,
            factura.iva,
            factura.total,
            factura.estado
        ])
    
    return response


# Add these imports at the top of the file if they don't exist already
from django.shortcuts import render
from django.core.paginator import Paginator
from apps.transacciones.models import Transaccion
from django.db.models import Sum, Q
from datetime import datetime, timedelta
from django.utils import timezone
import json

# Add or update the ventas view function
def ventas(request):
    # Get filter parameters
    fecha_desde = request.GET.get('fecha_desde', None)
    fecha_hasta = request.GET.get('fecha_hasta', None)
    tipo_documento = request.GET.get('tipo_documento', 'todos')
    
    # Set default date range if not provided (last 30 days)
    if not fecha_desde:
        fecha_desde = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_hasta:
        fecha_hasta = timezone.now().strftime('%Y-%m-%d')
    
    # Convert string dates to datetime objects
    fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
    fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
    fecha_hasta_dt = datetime.combine(fecha_hasta_dt.date(), datetime.max.time())  # Set to end of day
    
    # Query transactions based on filters
    transacciones = Transaccion.objects.filter(
        fecha__gte=fecha_desde_dt,
        fecha__lte=fecha_hasta_dt,
        procesado_pago=True,
        usuario_creador=request.user
    ).order_by('-fecha')
    
    # Apply document type filter if specified
    if tipo_documento != 'todos':
        # This is a simplified implementation - adjust based on your actual data model
        if tipo_documento == 'factura':
            # Filter for invoices
            transacciones = transacciones.filter(venta__tipo_documento='factura')
        elif tipo_documento == 'nota_venta':
            # Filter for sales notes
            transacciones = transacciones.filter(venta__tipo_documento='nota_venta')
        elif tipo_documento == 'ticket':
            # Filter for tickets
            transacciones = transacciones.filter(venta__tipo_documento='ticket')
    
    # Calculate summary statistics
    total_ventas = transacciones.aggregate(total=Sum('monto'))['total'] or 0
    
    # Calculate IVA correctly - moved after transacciones is defined
    total_iva = 0
    for transaccion in transacciones:
        # Calculate IVA as 15% of the subtotal (or total/1.15)
        total = float(transaccion.monto)
        subtotal = total / 1.15
        iva = total - subtotal
        total_iva += iva
    
    total_facturas = transacciones.count()
    
    # MEJORA 1: Calcular Ticket Promedio
    ticket_promedio = total_ventas / total_facturas if total_facturas > 0 else 0
    
    # Calculate percentage change compared to previous period
    previous_start = fecha_desde_dt - (fecha_hasta_dt - fecha_desde_dt)
    previous_end = fecha_desde_dt - timedelta(days=1)
    
    previous_ventas = Transaccion.objects.filter(
        fecha__gte=previous_start,
        fecha__lte=previous_end,
        procesado_pago=True,
        usuario_creador=request.user
    ).aggregate(total=Sum('monto'))['total'] or 0
    
    if previous_ventas > 0:
        porcentaje_cambio = ((total_ventas - previous_ventas) / previous_ventas) * 100
    else:
        porcentaje_cambio = 100 if total_ventas > 0 else 0
    
    # Prepare data for chart
    # Group transactions by date and sum amounts
    ventas_por_dia = {}
    for t in transacciones:
        fecha_str = t.fecha.strftime('%Y-%m-%d')
        if fecha_str in ventas_por_dia:
            ventas_por_dia[fecha_str] += float(t.monto)
        else:
            ventas_por_dia[fecha_str] = float(t.monto)
    
    # Sort by date and prepare for chart
    sorted_dates = sorted(ventas_por_dia.keys())
    ventas_chart_data = {
        'labels': [datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m') for d in sorted_dates],
        'values': [ventas_por_dia[d] for d in sorted_dates]
    }
    
    # Additional statistics for SRI (tax authority)
    total_ventas_gravadas = total_ventas
    # Adjust this calculation based on your actual data model if you have tax-exempt sales
    
    proxima_declaracion = (datetime.now().replace(day=1) + timedelta(days=32)).replace(day=1) + timedelta(days=9)
    retenciones_recibidas = 0  # You would calculate this from your data
    retenciones_efectuadas = 0  # You would calculate this from your data
    estado_ats = 'pendiente'  # This would be determined by your business logic
    
    # Pagination
    paginator = Paginator(transacciones, 10)  # Show 25 transactions per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Prepare period text for display
    days_diff = (fecha_hasta_dt - fecha_desde_dt).days
    if days_diff <= 1:
        periodo_texto = "Hoy"
    elif days_diff <= 7:
        periodo_texto = "Últimos 7 días"
    elif days_diff <= 30:
        periodo_texto = "Últimos 30 días"
    else:
        periodo_texto = f"Período de {days_diff} días"
    
    # Check if export was requested
    export_format = request.GET.get('export')
    if export_format in ['excel', 'pdf']:
        if export_format == 'excel':
            return exportar_excel(transacciones, fecha_desde_dt, fecha_hasta_dt, total_ventas, total_iva, total_facturas, ticket_promedio)
        elif export_format == 'pdf':
            return exportar_pdf(transacciones, fecha_desde_dt, fecha_hasta_dt, total_ventas, total_iva, total_facturas, ticket_promedio, request.user)
    
    context = {
        'transacciones': page_obj,
        'fecha_desde': fecha_desde_dt,
        'fecha_hasta': fecha_hasta_dt,
        'tipo_documento': tipo_documento,
        'total_ventas': total_ventas,
        'total_iva': total_iva,
        'total_facturas': total_facturas,
        'ticket_promedio': ticket_promedio,  # MEJORA 1: Agregar ticket promedio
        'porcentaje_cambio': porcentaje_cambio,
        'ventas_chart_data': json.dumps(ventas_chart_data),
        'periodo_texto': periodo_texto,
        'total_ventas_gravadas': total_ventas_gravadas,
        'proxima_declaracion': proxima_declaracion,
        'retenciones_recibidas': retenciones_recibidas,
        'retenciones_efectuadas': retenciones_efectuadas,
        'estado_ats': estado_ats,
        'page_obj': page_obj,
    }
    
    return render(request, 'reportes/ventas.html', context)

def iva_declaracion(request):
    """View for preparing IVA declaration"""
    # Get the current date range
    fecha_desde = request.GET.get('fecha_desde', None)
    fecha_hasta = request.GET.get('fecha_hasta', None)
    
    # If dates are provided, use them for the declaration
    if fecha_desde and fecha_hasta:
        # Fetch transactions for this period
        transacciones = Transaccion.objects.filter(
            fecha__gte=fecha_desde,
            fecha__lte=fecha_hasta
        )
        
        # Calculate totals
        total_ventas = sum(t.monto for t in transacciones)
        total_iva = 0
        for t in transacciones:
            total = float(t.monto)
            subtotal = total / 1.15
            iva = total - subtotal
            total_iva += iva
        
        context = {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'total_ventas': total_ventas,
            'total_iva': total_iva,
            'total_ventas_gravadas': total_ventas - total_iva,
            'transacciones': transacciones
        }
        
        return render(request, 'reportes/iva_declaracion.html', context)
    
    # If no dates, redirect back with a message
    messages.warning(request, 'Por favor seleccione un rango de fechas para preparar la declaración')
    return redirect('reportes:ventas')

def ats(request):
    """View for downloading ATS file in XML format according to SRI requirements"""
    # Get the current date range
    fecha_desde = request.GET.get('fecha_desde', None)
    fecha_hasta = request.GET.get('fecha_hasta', None)
    
    if not fecha_desde or not fecha_hasta:
        messages.warning(request, 'Por favor seleccione un rango de fechas para generar el ATS')
        return redirect('reportes:ventas')
    
    # Convert string dates to datetime objects
    fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
    fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
    fecha_hasta_dt = datetime.combine(fecha_hasta_dt.date(), datetime.max.time())  # Set to end of day
    
    # Fetch transactions for this period
    transacciones = Transaccion.objects.filter(
        fecha__gte=fecha_desde_dt,
        fecha__lte=fecha_hasta_dt,
        procesado_pago=True,
        usuario_creador=request.user
    )
    
    # Create XML structure
    import xml.dom.minidom as minidom
    import xml.etree.ElementTree as ET
    
    # Create root element
    root = ET.Element("iva")
    
    # Add header information
    cabecera = ET.SubElement(root, "TipoIDInformante")
    cabecera.text = "R"  # R for RUC
    
    # Get business info from the business configuration
    from apps.configuraciones.models import BusinessConfiguration
    
    # Try to get the business configuration
    try:
        business_config = BusinessConfiguration.objects.first()
        
        # Get RUC directly from business configuration and ensure it's a string
        if business_config and business_config.ruc_negocio:
            ruc_to_use = str(business_config.ruc_negocio).strip()
            razon_social = business_config.nombre_negocio
        else:
            ruc_to_use = '9999999999999'
            razon_social = getattr(request.user, 'razon_social', request.user.get_full_name())
    except Exception as e:
        # Log the error and use default values
        print(f"Error retrieving business config: {e}")
        ruc_to_use = '9999999999999'
        razon_social = getattr(request.user, 'razon_social', request.user.get_full_name())
    
    id_informante = ET.SubElement(root, "IdInformante")
    id_informante.text = ruc_to_use
    
    razon_social_elem = ET.SubElement(root, "razonSocial")
    razon_social_elem.text = razon_social
    
    anio = ET.SubElement(root, "Anio")
    anio.text = fecha_desde_dt.strftime('%Y')
    
    mes = ET.SubElement(root, "Mes")
    mes.text = fecha_desde_dt.strftime('%m')
    
    # Add sales information
    ventas = ET.SubElement(root, "comprobantesEmitidos")
    
    # Group transactions by date for summary
    ventas_por_dia = {}
    for t in transacciones:
        fecha_str = t.fecha.strftime('%Y-%m-%d')
        if fecha_str in ventas_por_dia:
            ventas_por_dia[fecha_str].append(t)
        else:
            ventas_por_dia[fecha_str] = [t]
    
    # Add each day's transactions
    for fecha, trans_list in ventas_por_dia.items():
        # Calculate totals for the day
        total_dia = sum(float(t.monto) for t in trans_list)
        subtotal_dia = total_dia / 1.15
        iva_dia = total_dia - subtotal_dia
        
        # Create daily summary
        dia = ET.SubElement(ventas, "dia")
        
        fecha_elem = ET.SubElement(dia, "fecha")
        fecha_elem.text = fecha
        
        # Add transaction details
        for t in trans_list:
            total = float(t.monto)
            subtotal = total / 1.15
            iva = total - subtotal
            
            # Handle cases where venta or cliente might be None
            cliente_nombre = "Cliente General"
            cliente_id = "9999999999"
            
            if hasattr(t, 'venta') and t.venta is not None:
                if hasattr(t.venta, 'cliente') and t.venta.cliente is not None:
                    cliente_nombre = t.venta.cliente.nombre
                    # Convert cliente_id to string to avoid TypeError
                    cliente_id = str(t.venta.cliente.identificacion or "9999999999")
            
            # Ensure factura_id is a string
            factura_id = str(getattr(t, 'factuID', '-') or '-')
            
            # Create transaction element
            transaccion = ET.SubElement(dia, "comprobante")
            
            tipo_doc = ET.SubElement(transaccion, "tipoComprobante")
            tipo_doc.text = "01"  # 01 for Factura
            
            serie = ET.SubElement(transaccion, "serie")
            serie.text = factura_id[:6] if len(factura_id) >= 6 else "001001"
            
            secuencial = ET.SubElement(transaccion, "secuencial")
            secuencial.text = factura_id[6:] if len(factura_id) >= 6 else str(t.pk).zfill(9)
            
            fecha_emision = ET.SubElement(transaccion, "fechaEmision")
            fecha_emision.text = t.fecha.strftime('%d/%m/%Y')
            
            # Cliente information
            cliente = ET.SubElement(transaccion, "cliente")
            
            tipo_id_cliente = ET.SubElement(cliente, "tipoIdentificacion")
            # Ensure cliente_id is a string before checking length
            cliente_id_str = str(cliente_id).strip()
            tipo_id_cliente.text = "04" if len(cliente_id_str) == 10 else "01"
            
            id_cliente = ET.SubElement(cliente, "identificacion")
            id_cliente.text = cliente_id_str
            
            nombre_cliente = ET.SubElement(cliente, "razonSocial")
            nombre_cliente.text = cliente_nombre
            
            # Values
            valores = ET.SubElement(transaccion, "valores")
            
            base_imponible = ET.SubElement(valores, "baseImponible")
            base_imponible.text = f"{subtotal:.2f}"
            
            impuesto = ET.SubElement(valores, "impuesto")
            impuesto.text = f"{iva:.2f}"
            
            total_elem = ET.SubElement(valores, "total")
            total_elem.text = f"{total:.2f}"
    
    # Create pretty XML string
    xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="   ")
    
    # Create response
    response = HttpResponse(xml_str, content_type='application/xml')
    response['Content-Disposition'] = f'attachment; filename=ATS_{datetime.now().strftime("%Y%m%d")}.xml'
    
    return response

# Funciones de exportación
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

def exportar_excel(transacciones, fecha_desde, fecha_hasta, total_ventas, total_iva, total_facturas, ticket_promedio):
    """Exportar reporte de ventas a Excel - Formato profesional para contadores"""
    
    # Convertir Decimal a float al inicio para evitar errores de tipo
    total_ventas = float(total_ventas)
    total_iva = float(total_iva)
    ticket_promedio = float(ticket_promedio)
    
    wb = Workbook()
    
    # === HOJA 1: RESUMEN EJECUTIVO ===
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # Encabezado principal
    ws_resumen['A1'] = 'LEMON POS - REPORTE DE VENTAS'
    ws_resumen['A1'].font = Font(size=18, bold=True, color="22C55E")
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    ws_resumen.merge_cells('A1:B1')
    
    # Período
    ws_resumen['A3'] = 'Período:'
    ws_resumen['B3'] = f"{fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    ws_resumen['A3'].font = Font(bold=True)
    
    ws_resumen['A4'] = 'Fecha de generación:'
    ws_resumen['B4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    # KPIs principales
    ws_resumen['A6'] = 'INDICADORES CLAVE'
    ws_resumen['A6'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_resumen['A6'].fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    ws_resumen['A6'].alignment = Alignment(horizontal='center')
    ws_resumen.merge_cells('A6:B6')
    
    kpis = [
        ('Ventas Totales', f"${total_ventas:.2f}"),
        ('Facturas Emitidas', total_facturas),
        ('Ticket Promedio', f"${ticket_promedio:.2f}"),
        ('IVA Generado (15%)', f"${total_iva:.2f}"),
        ('Ventas Gravadas', f"${(total_ventas - total_iva):.2f}")
    ]
    
    row = 7
    for label, value in kpis:
        ws_resumen[f'A{row}'] = label
        ws_resumen[f'B{row}'] = value
        ws_resumen[f'A{row}'].font = Font(bold=True)
        if row == 7:  # Ventas totales destacadas
            ws_resumen[f'B{row}'].font = Font(size=14, bold=True, color="22C55E")
        row += 1
    
    # MEJORA 4: Distribución por Método de Pago
    ws_resumen[f'A{row+1}'] = 'DISTRIBUCIÓN POR MÉTODO DE PAGO'
    ws_resumen[f'A{row+1}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_resumen[f'A{row+1}'].fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    ws_resumen[f'A{row+1}'].alignment = Alignment(horizontal='center')
    ws_resumen.merge_cells(f'A{row+1}:B{row+1}')
    
    # Calcular totales por método
    metodos_pago = {}
    for t in transacciones:
        metodo = t.metodo_pago
        if metodo == 'cash':
            metodo_es = 'Efectivo'
        elif metodo == 'card':
            metodo_es = 'Tarjeta'
        elif metodo == 'transfer':
            metodo_es = 'Transferencia'
        else:
            metodo_es = metodo.title()
        
        if metodo_es not in metodos_pago:
            metodos_pago[metodo_es] = 0
        metodos_pago[metodo_es] += float(t.monto)
    
    row += 3
    for metodo, monto in metodos_pago.items():
        ws_resumen[f'A{row}'] = metodo
        ws_resumen[f'B{row}'] = f"${monto:.2f}"
        ws_resumen[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['B'].width = 20
    
    # === HOJA 2: DETALLE COMPLETO (Para contador) ===
    ws_detalle = wb.create_sheet("Detalle Completo")
    
    # Encabezados técnicos
    headers = ['ID Transacción', 'Factura #', 'Fecha', 'Hora', 'Cliente', 'Subtotal', 'IVA 15%', 'Total', 'Método Pago', 'Estado']
    ws_detalle.append(headers)
    
    # Estilo de encabezados
    for idx, cell in enumerate(ws_detalle[1], 1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Congelar fila de encabezados (MEJORA 1)
    ws_detalle.freeze_panes = 'A2'
    
    # Datos completos
    for t in transacciones:
        subtotal = float(t.monto) / 1.15
        iva = float(t.monto) - subtotal
        
        # Traducir método de pago a español
        metodo = t.metodo_pago
        if metodo == 'cash':
            metodo_es = 'Efectivo'
        elif metodo == 'card':
            metodo_es = 'Tarjeta'
        elif metodo == 'transfer':
            metodo_es = 'Transferencia'
        else:
            metodo_es = metodo.title()
        
        # Remover timezone para Excel
        fecha_sin_tz = t.fecha.replace(tzinfo=None) if t.fecha.tzinfo else t.fecha
        
        ws_detalle.append([
            t.transaction_id,
            t.factuID,
            fecha_sin_tz,  # Fecha sin timezone
            fecha_sin_tz,  # Hora sin timezone
            t.venta.cliente.nombre if t.venta.cliente else 'Consumidor Final',
            round(subtotal, 2),
            round(iva, 2),
            float(t.monto),
            metodo_es,  # Método en español
            'Pagado'
        ])
    
    # MEJORA 2: Formato correcto de columnas
    # Formato de fecha (columna C)
    for row in ws_detalle.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = 'DD/MM/YYYY'
    
    # Formato de hora (columna D)
    for row in ws_detalle.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = 'HH:MM'
    
    # Formato de moneda (columnas F, G, H)
    for row in ws_detalle.iter_rows(min_row=2, min_col=6, max_col=8):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    # Ajustar anchos
    ws_detalle.column_dimensions['A'].width = 15
    ws_detalle.column_dimensions['B'].width = 12
    ws_detalle.column_dimensions['C'].width = 12
    ws_detalle.column_dimensions['D'].width = 8
    ws_detalle.column_dimensions['E'].width = 30
    ws_detalle.column_dimensions['F'].width = 12
    ws_detalle.column_dimensions['G'].width = 12
    ws_detalle.column_dimensions['H'].width = 12
    ws_detalle.column_dimensions['I'].width = 15
    ws_detalle.column_dimensions['J'].width = 10
    
    # === HOJA 3: RESUMEN POR DÍA (Premium) ===
    ws_diario = wb.create_sheet("Resumen por Día")
    
    # Encabezados
    headers_diario = ['Fecha', 'Total Vendido', 'Número de Facturas', 'Ticket Promedio']
    ws_diario.append(headers_diario)
    
    for cell in ws_diario[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Agrupar por día
    ventas_por_dia = {}
    for t in transacciones:
        fecha_str = t.fecha.strftime('%Y-%m-%d')
        if fecha_str not in ventas_por_dia:
            ventas_por_dia[fecha_str] = {'total': 0, 'facturas': 0}
        ventas_por_dia[fecha_str]['total'] += float(t.monto)
        ventas_por_dia[fecha_str]['facturas'] += 1
    
    # Agregar datos
    for fecha_str in sorted(ventas_por_dia.keys()):
        data = ventas_por_dia[fecha_str]
        ticket_prom = data['total'] / data['facturas'] if data['facturas'] > 0 else 0
        ws_diario.append([
            datetime.strptime(fecha_str, '%Y-%m-%d').strftime('%d/%m/%Y'),
            round(data['total'], 2),
            data['facturas'],
            round(ticket_prom, 2)
        ])
    
    # Formato
    for row in ws_diario.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '$#,##0.00'
    for row in ws_diario.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    # Ajustar anchos
    ws_diario.column_dimensions['A'].width = 15
    ws_diario.column_dimensions['B'].width = 15
    ws_diario.column_dimensions['C'].width = 20
    ws_diario.column_dimensions['D'].width = 18
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ventas_lemon_{fecha_desde.strftime("%Y%m%d")}_{fecha_hasta.strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response

def exportar_pdf(transacciones, fecha_desde, fecha_hasta, total_ventas, total_iva, total_facturas, ticket_promedio, user):
    """Exportar reporte de ventas a PDF - Formato ejecutivo y presentable"""
    
    # Convertir Decimal a float al inicio para evitar errores de tipo
    total_ventas = float(total_ventas)
    total_iva = float(total_iva)
    ticket_promedio = float(ticket_promedio)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=reporte_ventas_lemon_{fecha_desde.strftime("%Y%m%d")}_{fecha_hasta.strftime("%Y%m%d")}.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # === PORTADA / ENCABEZADO ===
    # Logo y título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#22C55E'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph("🍋 LEMON POS", title_style))
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.HexColor('#666666'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Reporte de Ventas", subtitle_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Información del negocio
    if hasattr(user, 'business') and user.business:
        business_data = [
            ['Negocio:', user.business.nombre_negocio or 'N/A'],
            ['RUC:', user.business.ruc_negocio or 'N/A'],
            ['Sucursal:', 'Matriz']
        ]
        business_table = Table(business_data, colWidths=[1.5*inch, 4*inch])
        business_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(business_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Período y fecha de generación
    periodo_data = [
        ['Período:', f"{fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"],
        ['Generado:', datetime.now().strftime('%d/%m/%Y %H:%M')]
    ]
    periodo_table = Table(periodo_data, colWidths=[1.5*inch, 4*inch])
    periodo_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(periodo_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # === RESUMEN EJECUTIVO ===
    elements.append(Paragraph("<b>RESUMEN EJECUTIVO</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.2*inch))
    
    # KPIs en tabla destacada
    kpi_data = [
        ['Ventas Totales', f'${float(total_ventas):,.2f}'],
        ['Facturas Emitidas', f'{total_facturas:,}'],
        ['Ticket Promedio', f'${float(ticket_promedio):,.2f}'],
        ['IVA Generado', f'${float(total_iva):,.2f}']
    ]
    
    kpi_table = Table(kpi_data, colWidths=[3.5*inch, 2*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0FDF4')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#166534')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#22C55E')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('FONTSIZE', (1, 0), (1, 0), 16),  # Ventas totales más grande
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#22C55E')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#22C55E')),
    ]))
    
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # === DISTRIBUCIÓN POR MÉTODO DE PAGO (Premium) ===
    elements.append(Paragraph("<b>DISTRIBUCIÓN POR MÉTODO DE PAGO</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.15*inch))
    
    # Calcular totales por método
    metodos_pago = {}
    for t in transacciones:
        metodo = t.metodo_pago
        if metodo == 'cash':
            metodo_es = 'Efectivo'
        elif metodo == 'card':
            metodo_es = 'Tarjeta'
        elif metodo == 'transfer':
            metodo_es = 'Transferencia'
        else:
            metodo_es = metodo.title()
        
        if metodo_es not in metodos_pago:
            metodos_pago[metodo_es] = 0
        metodos_pago[metodo_es] += float(t.monto)
    
    metodos_data = [[metodo, f'${monto:,.2f}'] for metodo, monto in metodos_pago.items()]
    metodos_table = Table(metodos_data, colWidths=[2.5*inch, 1.5*inch])
    metodos_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0FDF4')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#22C55E')),
    ]))
    elements.append(metodos_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # === DETALLE DE TRANSACCIONES ===
    elements.append(Paragraph("<b>DETALLE DE TRANSACCIONES</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Tabla mejorada con hora y mejor formato
    trans_data = [['Factura #', 'Fecha', 'Hora', 'Cliente', 'Total', 'Método']]
    
    for t in transacciones[:50]:
        cliente = t.venta.cliente.nombre if t.venta.cliente else 'Consumidor Final'
        if len(cliente) > 22:
            cliente = cliente[:19] + '...'
        
        # Traducir método de pago
        metodo = t.metodo_pago
        if metodo == 'cash':
            metodo_es = 'Efectivo'
        elif metodo == 'card':
            metodo_es = 'Tarjeta'
        elif metodo == 'transfer':
            metodo_es = 'Transferencia'
        else:
            metodo_es = metodo.title()
        
        trans_data.append([
            str(t.factuID),
            t.fecha.strftime('%d/%m/%Y'),
            t.fecha.strftime('%H:%M'),
            cliente,
            f'${t.monto:,.2f}',
            metodo_es
        ])
    
    trans_table = Table(trans_data, colWidths=[0.8*inch, 0.9*inch, 0.6*inch, 2*inch, 0.9*inch, 1*inch])
    trans_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22C55E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),  # Total alineado a la derecha
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 7.5),  # Texto más pequeño
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])  # Zebra stripes
    ]))
    
    elements.append(trans_table)
    
    # === RESUMEN FINAL DE VERIFICACIÓN ===
    elements.append(Spacer(1, 0.3*inch))
    verificacion_data = [
        ['Total de registros:', f'{transacciones.count():,}'],
        ['Total verificado:', f'${total_ventas:,.2f}']
    ]
    verificacion_table = Table(verificacion_data, colWidths=[3*inch, 2*inch])
    verificacion_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#22C55E')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEABOVE', (0, 0), (-1, 0), 1, colors.grey),
    ]))
    elements.append(verificacion_table)
    
    # Nota si hay más registros
    if transacciones.count() > 50:
        elements.append(Spacer(1, 0.2*inch))
        note_style = ParagraphStyle(
            'Note',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#666666'),
            alignment=TA_CENTER,
            fontName='Helvetica-Oblique'
        )
        elements.append(Paragraph(f"Mostrando 50 de {transacciones.count()} transacciones. Descargue el Excel para ver el detalle completo.", note_style))
    
    # === RESUMEN TRIBUTARIO ===
    elements.append(Spacer(1, 0.4*inch))
    elements.append(Paragraph("<b>RESUMEN TRIBUTARIO</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.2*inch))
    
    ventas_gravadas = float(total_ventas) - float(total_iva)
    sri_data = [
        ['Ventas Gravadas:', f'${ventas_gravadas:,.2f}'],
        ['IVA Generado (15%):', f'${float(total_iva):,.2f}'],
        ['Estado SRI:', 'Pendiente de declaración']
    ]
    
    sri_table = Table(sri_data, colWidths=[3*inch, 2.5*inch])
    sri_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor('#FCD34D')),
    ]))
    
    elements.append(sri_table)
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#999999'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Generado por Lemon POS - Sistema de Gestión Empresarial", footer_style))
    
    doc.build(elements)
    return response


# ============================================
# FUNCIONES DE EXPORTACIÓN DE INVENTARIO
# ============================================

def exportar_inventario_excel(productos, categorias, valor_total_costo, valor_total_venta, 
                               productos_stock_bajo, productos_agotados, valor_en_riesgo, 
                               utilidad_potencial, promedio_valor, user):
    """Exportar inventario a Excel - Formato profesional multi-hoja"""
    from decimal import Decimal
    from apps.usuarios.models import Business
    
    # Convertir Decimal a float
    valor_total_costo = float(valor_total_costo)
    valor_total_venta = float(valor_total_venta)
    valor_en_riesgo = float(valor_en_riesgo)
    utilidad_potencial = float(utilidad_potencial)
    promedio_valor = float(promedio_valor)
    
    # 🎯 Obtener información del negocio del usuario
    try:
        business = Business.objects.get(user=user)
        nombre_negocio = business.nombre_negocio or user.nombre_completo
        ruc_negocio = str(business.ruc_negocio) if business.ruc_negocio else 'N/A'
    except Business.DoesNotExist:
        nombre_negocio = user.nombre_completo
        ruc_negocio = user.ruc_negocio if hasattr(user, 'ruc_negocio') else 'N/A'
    except Exception as e:
        nombre_negocio = 'Lemon POS'
        ruc_negocio = 'N/A'
    
    wb = Workbook()
    
    # === HOJA 1: RESUMEN EJECUTIVO ===
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # ENCABEZADO FORMAL
    ws_resumen['A1'] = 'REPORTE DE INVENTARIO'
    ws_resumen['A1'].font = Font(size=18, bold=True, color="22C55E")
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    ws_resumen.merge_cells('A1:B1')
    
    # Información del negocio
    ws_resumen['A3'] = 'Nombre del negocio:'
    ws_resumen['B3'] = nombre_negocio
    ws_resumen['A3'].font = Font(bold=True)
    
    ws_resumen['A4'] = 'RUC:'
    ws_resumen['B4'] = ruc_negocio
    ws_resumen['A4'].font = Font(bold=True)
    
    ws_resumen['A5'] = 'Fecha de generación:'
    ws_resumen['B5'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws_resumen['A5'].font = Font(bold=True)
    
    ws_resumen['A6'] = 'Ambiente:'
    ws_resumen['B6'] = 'Producción'
    ws_resumen['A6'].font = Font(bold=True)
    
    # KPIs principales
    ws_resumen['A8'] = 'RESUMEN DEL INVENTARIO'
    ws_resumen['A8'].font = Font(size=12, bold=True, color="FFFFFF")
    ws_resumen['A8'].fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    ws_resumen['A8'].alignment = Alignment(horizontal='center')
    ws_resumen.merge_cells('A8:B8')
    
    kpis = [
        ('Total Productos', productos.count()),
        ('Valor Total Inventario (Costo)', f"${valor_total_costo:.2f}"),
        ('Valor Total Inventario (Venta)', f"${valor_total_venta:.2f}"),
        ('Utilidad Potencial', f"${utilidad_potencial:.2f}"),
        ('Promedio Valor por Producto', f"${promedio_valor:.2f}"),
        ('Productos con Stock Bajo', productos_stock_bajo.count()),
        ('Productos Agotados', productos_agotados.count()),
        ('Valor en Riesgo (Stock Bajo)', f"${valor_en_riesgo:.2f}"),
    ]
    
    row = 9
    for label, value in kpis:
        ws_resumen[f'A{row}'] = label
        ws_resumen[f'B{row}'] = value
        ws_resumen[f'A{row}'].font = Font(bold=True)
        if row == 10:  # Valor total destacado
            ws_resumen[f'B{row}'].font = Font(size=14, bold=True, color="22C55E")
        row += 1
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 25
    
    # === HOJA 2: DETALLE COMPLETO ===
    ws_detalle = wb.create_sheet("Detalle Completo")
    
    # 🎯 Encabezados profesionales con nuevos campos
    headers = ['Producto', 'SKU', 'Tipo', 'Categoría', 'Stock', 'Stock Mín', 'Unidad', 'Costo Unit.', 'Precio Venta', 'Valor Costo', 'Valor Venta', 'Margen %']
    ws_detalle.append(headers)
    
    # Estilo de encabezados
    for idx, cell in enumerate(ws_detalle[1], 1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Congelar fila de encabezados
    ws_detalle.freeze_panes = 'A2'
    
    # Datos completos
    for p in productos:
        if p.stock is not None:
            costo = float(p.costo) if p.costo else 0
            precio_base = float(p.precio_base)
            precio_venta = float(p.precio)
            stock = p.stock or 0
            stock_minimo = p.stock_minimo or 0
            valor_costo = costo * stock
            valor_venta = precio_venta * stock
            margen_porcentaje = float(p.margen_utilidad) if p.costo and p.costo > 0 else 0
            
            # Traducir tipo de producto
            tipo_display = dict(p.TIPO_PRODUCTO_CHOICES).get(p.tipo_producto, p.tipo_producto)
            unidad_display = dict(p.UNIDAD_MEDIDA_CHOICES).get(p.unidad_medida, p.unidad_medida)
            
            ws_detalle.append([
                p.nombre,
                p.sku if p.sku else 'N/A',
                tipo_display,
                p.categoria.nombre if p.categoria else 'Sin categoría',
                stock,
                stock_minimo,
                unidad_display,
                costo,
                precio_venta,
                valor_costo,
                valor_venta,
                margen_porcentaje
            ])
    
    # Formato de moneda (columnas H, I, J, K)
    for row in ws_detalle.iter_rows(min_row=2, min_col=8, max_col=11):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    # Formato de porcentaje (columna L)
    for row in ws_detalle.iter_rows(min_row=2, min_col=12, max_col=12):
        for cell in row:
            cell.number_format = '0.00"%"'
    
    # Ajustar anchos
    ws_detalle.column_dimensions['A'].width = 30
    ws_detalle.column_dimensions['B'].width = 15
    ws_detalle.column_dimensions['C'].width = 18
    ws_detalle.column_dimensions['D'].width = 20
    ws_detalle.column_dimensions['E'].width = 10
    ws_detalle.column_dimensions['F'].width = 12
    ws_detalle.column_dimensions['G'].width = 12
    ws_detalle.column_dimensions['H'].width = 12
    ws_detalle.column_dimensions['I'].width = 13
    ws_detalle.column_dimensions['J'].width = 14
    ws_detalle.column_dimensions['K'].width = 14
    ws_detalle.column_dimensions['L'].width = 12
    
    # === HOJA 3: PRODUCTOS CON STOCK BAJO ===
    ws_bajo = wb.create_sheet("Stock Bajo")
    
    headers_bajo = ['Producto', 'Categoría', 'Stock Actual', 'Precio', 'Valor', 'Estado']
    ws_bajo.append(headers_bajo)
    
    for cell in ws_bajo[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for p in productos_stock_bajo:
        ws_bajo.append([
            p.nombre,
            p.categoria.nombre if p.categoria else 'Sin categoría',
            p.stock,
            float(p.precio),
            float(p.precio) * (p.stock or 0),
            'CRÍTICO' if p.stock < 5 else 'BAJO'
        ])
    
    # Formato
    for row in ws_bajo.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    ws_bajo.column_dimensions['A'].width = 30
    ws_bajo.column_dimensions['B'].width = 20
    ws_bajo.column_dimensions['C'].width = 12
    ws_bajo.column_dimensions['D'].width = 12
    ws_bajo.column_dimensions['E'].width = 15
    ws_bajo.column_dimensions['F'].width = 12
    
    # === HOJA 4: RESUMEN POR CATEGORÍA ===
    ws_categorias = wb.create_sheet("Por Categoría")
    
    headers_cat = ['Categoría', 'Total Productos', 'Valor Inventario (Costo)', 'Valor Inventario (Venta)', 'Utilidad Potencial']
    ws_categorias.append(headers_cat)
    
    for cell in ws_categorias[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for cat in categorias:
        valor_costo = float(cat.valor_inventario_costo)
        valor_venta = float(cat.valor_inventario_venta)
        utilidad = valor_venta - valor_costo
        
        ws_categorias.append([
            cat.nombre,
            cat.total_productos,
            valor_costo,
            valor_venta,
            utilidad
        ])
    
    # Formato
    for row in ws_categorias.iter_rows(min_row=2, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    ws_categorias.column_dimensions['A'].width = 25
    ws_categorias.column_dimensions['B'].width = 18
    ws_categorias.column_dimensions['C'].width = 22
    ws_categorias.column_dimensions['D'].width = 22
    ws_categorias.column_dimensions['E'].width = 20
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=inventario_lemon_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response


def exportar_inventario_pdf(productos, categorias, valor_total_costo, valor_total_venta,
                             productos_stock_bajo, productos_agotados, valor_en_riesgo,
                             utilidad_potencial, user):
    """Exportar inventario a PDF - Formato ejecutivo"""
    from decimal import Decimal
    from apps.usuarios.models import Business
    
    # Convertir Decimal a float
    valor_total_costo = float(valor_total_costo)
    valor_total_venta = float(valor_total_venta)
    valor_en_riesgo = float(valor_en_riesgo)
    utilidad_potencial = float(utilidad_potencial)
    
    # 🎯 Obtener información del negocio del usuario
    try:
        business = Business.objects.get(user=user)
        nombre_negocio = business.nombre_negocio or user.nombre_completo
        ruc_negocio = str(business.ruc_negocio) if business.ruc_negocio else 'N/A'
    except Business.DoesNotExist:
        nombre_negocio = user.nombre_completo
        ruc_negocio = user.ruc_negocio if hasattr(user, 'ruc_negocio') else 'N/A'
    except Exception as e:
        nombre_negocio = 'Lemon POS'
        ruc_negocio = 'N/A'
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=inventario_{nombre_negocio.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    
    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#22C55E'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Título
    elements.append(Paragraph("REPORTE DE INVENTARIO", title_style))
    
    # ENCABEZADO FORMAL
    header_data = [
        ['Nombre del negocio:', nombre_negocio],
        ['RUC:', ruc_negocio],
        ['Fecha de generación:', datetime.now().strftime('%d/%m/%Y %H:%M')],
        ['Ambiente:', 'Producción'],
    ]
    
    header_table = Table(header_data, colWidths=[2*inch, 4*inch])
    header_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # === RESUMEN EJECUTIVO ===
    resumen_data = [
        ['RESUMEN DEL INVENTARIO', ''],
        ['Total Productos', str(productos.count())],
        ['Valor Total Inventario (Costo)', f'${valor_total_costo:,.2f}'],
        ['Valor Total Inventario (Venta)', f'${valor_total_venta:,.2f}'],
        ['Utilidad Potencial', f'${utilidad_potencial:,.2f}'],
        ['Productos con Stock Bajo', str(productos_stock_bajo.count())],
        ['Productos Agotados', str(productos_agotados.count())],
        ['Valor en Riesgo (Stock Bajo)', f'${valor_en_riesgo:,.2f}'],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[4*inch, 2*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22C55E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    elements.append(resumen_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # === DETALLE DE PRODUCTOS ===
    elements.append(Paragraph("DETALLE DE PRODUCTOS", styles['Heading2']))
    elements.append(Spacer(1, 0.2*inch))
    
    # 🎯 Encabezados de tabla profesional con nuevos campos
    detalle_data = [['Producto', 'SKU', 'Tipo', 'Categoría', 'Stock', 'Stock Mín', 'Unidad', 'Costo', 'Precio', 'Valor Costo', 'Valor Venta', 'Margen %']]
    
    # Agregar productos (limitar a primeros 25 para no saturar el PDF)
    count = 0
    for p in productos:
        if p.stock is not None and count < 25:
            costo = float(p.costo) if p.costo else 0
            precio = float(p.precio)
            stock = p.stock or 0
            stock_minimo = p.stock_minimo or 0
            valor_costo = costo * stock
            valor_venta = precio * stock
            margen = float(p.margen_utilidad) if p.costo and p.costo > 0 else 0
            
            # Traducir tipo
            tipo_display = 'Físico' if p.tipo_producto == 'fisico' else 'Insumo'
            unidad_display = dict(p.UNIDAD_MEDIDA_CHOICES).get(p.unidad_medida, p.unidad_medida)[:3]
            
            detalle_data.append([
                p.nombre[:20],  # Truncar nombre largo
                (p.sku if p.sku else 'N/A')[:10],
                tipo_display[:6],
                (p.categoria.nombre if p.categoria else 'N/A')[:12],
                str(stock),
                str(stock_minimo),
                unidad_display,
                f'${costo:.2f}',
                f'${precio:.2f}',
                f'${valor_costo:.2f}',
                f'${valor_venta:.2f}',
                f'{margen:.1f}%'
            ])
            count += 1
    
    if productos.count() > 25:
        detalle_data.append(['...', '...', '...', '...', '...', '...', '...', '...', '...', '...', '...', '...'])
        detalle_data.append([f'Total: {productos.count()} productos', '', '', '', '', '', '', '', '', '', '', ''])
    
    detalle_table = Table(detalle_data, colWidths=[1.3*inch, 0.6*inch, 0.5*inch, 0.7*inch, 0.4*inch, 0.5*inch, 0.4*inch, 0.5*inch, 0.5*inch, 0.7*inch, 0.7*inch, 0.5*inch])
    detalle_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22C55E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (4, 0), (5, -1), 'CENTER'),
        ('ALIGN', (7, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(detalle_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # === TOTAL GENERAL ===
    total_data = [
        ['TOTAL GENERAL INVENTARIO', f'${valor_total_costo:,.2f}']
    ]
    
    total_table = Table(total_data, colWidths=[4*inch, 2*inch])
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22C55E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    ]))
    
    elements.append(total_table)
    
    # Construir PDF
    doc.build(elements)
    return response
